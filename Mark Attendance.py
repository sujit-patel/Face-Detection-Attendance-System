import cv2
import os
import openpyxl
from datetime import datetime

# Load pre-trained face detection model
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

# Initialize webcam
cap = cv2.VideoCapture(0)

# Directory where images are saved
img_dir = "img"

# Create or load Excel workbook
excel_file = "attendance.xlsx"
if not os.path.exists(excel_file):
    wb = openpyxl.Workbook()
else:
    wb = openpyxl.load_workbook(excel_file)

# Ask user to enter subject name
subject_name = input("Enter subject name: ")

# Create or load worksheet for the subject
if subject_name in wb.sheetnames:
    ws = wb[subject_name]
else:
    ws = wb.create_sheet(title=subject_name)
    ws.append(["Date", "Time", "ID", "Attendance"])

# Load saved face images and their corresponding IDs
saved_faces = {}
for img_name in os.listdir(img_dir):
    img_path = os.path.join(img_dir, img_name)
    if os.path.isfile(img_path):
        img_id, _ = os.path.splitext(img_name)
        img_id = img_id.lower()  # Convert to lowercase to match filename case
        img_id = img_id.replace(" ", "_")  # Replace spaces with underscores
        saved_faces[img_id] = cv2.imread(img_path, cv2.IMREAD_GRAYSCALE)

# List of IDs already marked for attendance
marked_ids = []

while True:
    # Read frame from webcam
    ret, frame = cap.read()

    # Convert frame to grayscale
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

    # Detect faces in the grayscale frame
    faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))

    # Draw rectangles around the detected faces and mark attendance
    for (x, y, w, h) in faces:
        face_detected = False
        for img_id, saved_face in saved_faces.items():
            # Compare detected face with saved faces
            result = cv2.matchTemplate(gray[y:y+h, x:x+w], saved_face, cv2.TM_CCOEFF_NORMED)
            _, confidence, _, _ = cv2.minMaxLoc(result)
            if confidence > 0.7:  # Threshold for confidence
                if img_id not in marked_ids:
                    current_time = datetime.now().strftime("%H:%M:%S")
                    current_date = datetime.now().strftime("%Y-%m-%d")
                    ws.append([current_date, current_time, img_id, "Present"])
                    marked_ids.append(img_id)
                    print("Attendance marked for ID:", img_id)
                    cv2.putText(frame, f"ID: {img_id}", (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)
                    cv2.putText(frame, "Attendance marked", (x, y-40), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)
                else:
                    print("Attendance already marked for ID:", img_id)
                    cv2.putText(frame, f"ID: {img_id}", (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)
                    cv2.putText(frame, "Attendance already marked", (x, y-40), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 0, 255), 2)
                face_detected = True
                break
        if not face_detected:
            cv2.putText(frame, "Save your face", (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 0, 255), 2)
        cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 255, 0), 2)

    # Display the frame with detected faces
    cv2.imshow('Face Detection', frame)

    # Exit loop if 'q' is pressed
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Mark absent for IDs not detected
current_time = datetime.now().strftime("%H:%M:%S")
current_date = datetime.now().strftime("%Y-%m-%d")
for img_id in saved_faces.keys():
    if img_id not in marked_ids:
        ws.append([current_date, current_time, img_id, "Absent"])
        print("Attendance marked as absent for ID:", img_id)

# Save the Excel workbook
wb.save(excel_file)
print("Attendance saved successfully!")

# Release the webcam and close all windows
cap.release()
cv2.destroyAllWindows()
